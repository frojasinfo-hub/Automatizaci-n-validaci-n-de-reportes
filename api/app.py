"""FastAPI backend para AutomatizaciónV1."""

from __future__ import annotations

import subprocess
import sys
from pathlib import Path
from typing import Any

import openpyxl
from fastapi import FastAPI, HTTPException, Query, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse

ROOT = Path(__file__).parent.parent
INPUT_PDFS = ROOT / "input" / "pdfs"
LOGS_DIR = ROOT / "logs"
REPORTS_FILE = ROOT / "reports" / "resultado_consultas.xlsx"
OUTPUT_CONSULTAS = ROOT / "output" / "CONSULTAS_AUTOMATIZADAS"

app = FastAPI(title="AutomatizaciónV1 API", version="1.0.0")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

_proc: subprocess.Popen | None = None  # type: ignore[type-arg]


# ── Upload ────────────────────────────────────────────────────────────────────

@app.post("/api/files/upload")
async def upload_file(file: UploadFile) -> dict[str, str]:
    if not file.filename or not file.filename.lower().endswith(".pdf"):
        raise HTTPException(status_code=400, detail="Solo se aceptan archivos PDF.")
    INPUT_PDFS.mkdir(parents=True, exist_ok=True)
    dest = INPUT_PDFS / file.filename
    dest.write_bytes(await file.read())
    return {"status": "ok", "filename": file.filename}


@app.get("/api/files/list")
async def list_files() -> list[dict[str, Any]]:
    if not INPUT_PDFS.exists():
        return []
    return [
        {"name": f.name, "size_kb": round(f.stat().st_size / 1024, 1)}
        for f in sorted(
            INPUT_PDFS.glob("*.pdf"),
            key=lambda p: p.stat().st_mtime,
            reverse=True,
        )
    ]


# ── Automation ────────────────────────────────────────────────────────────────

@app.post("/api/automation/run")
async def automation_run() -> dict[str, Any]:
    global _proc
    if _proc is not None and _proc.poll() is None:
        return {"status": "already_running", "pid": _proc.pid}

    # CREATE_NEW_CONSOLE abre ventana separada para que el operador resuelva CAPTCHAs
    _proc = subprocess.Popen(
        [sys.executable, "src/main.py"],
        cwd=str(ROOT),
        creationflags=subprocess.CREATE_NEW_CONSOLE,
    )
    return {"status": "started", "pid": _proc.pid}


@app.get("/api/automation/status")
async def automation_status() -> dict[str, Any]:
    global _proc
    if _proc is None:
        status = "idle"
    elif _proc.poll() is None:
        status = "running"
    elif _proc.returncode == 0:
        status = "done"
    else:
        status = "error"

    log_lines: list[str] = []
    log_file = LOGS_DIR / "ejecucion.log"
    if log_file.exists():
        try:
            text = log_file.read_text(encoding="utf-8", errors="replace")
            log_lines = text.splitlines()[-25:]
        except Exception:
            pass

    return {
        "status": status,
        "returncode": _proc.returncode if _proc else None,
        "log": log_lines,
    }


# ── Reports ───────────────────────────────────────────────────────────────────

@app.get("/api/reports/data")
async def reports_data() -> list[dict[str, Any]]:
    if not REPORTS_FILE.exists():
        return []
    try:
        wb = openpyxl.load_workbook(REPORTS_FILE, read_only=True, data_only=True)
        ws = wb.active
        headers = [
            str(cell.value) if cell.value is not None else f"Col{i}"
            for i, cell in enumerate(ws[1], 1)  # type: ignore[arg-type]
        ]
        rows: list[dict[str, Any]] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(v is not None for v in row):
                rows.append(
                    {h: (str(v) if v is not None else "") for h, v in zip(headers, row)}
                )
        wb.close()
        return rows
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc


@app.get("/api/reports/download")
async def reports_download() -> FileResponse:
    if not REPORTS_FILE.exists():
        raise HTTPException(status_code=404, detail="Reporte no encontrado.")
    return FileResponse(
        path=str(REPORTS_FILE),
        filename="resultado_consultas.xlsx",
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
    )


# ── Tercero PDFs ──────────────────────────────────────────────────────────────

@app.get("/api/tercero/pdfs")
async def list_tercero_pdfs(folder: str = Query(...)) -> list[dict[str, Any]]:
    carpeta = OUTPUT_CONSULTAS / folder
    if not carpeta.exists():
        return []
    try:
        carpeta.resolve().relative_to(OUTPUT_CONSULTAS.resolve())
    except ValueError:
        raise HTTPException(status_code=400, detail="Ruta no permitida.")
    return [
        {"name": f.name, "size_kb": round(f.stat().st_size / 1024, 1)}
        for f in sorted(carpeta.glob("*.pdf"))
    ]


@app.get("/api/tercero/pdf")
async def download_tercero_pdf(
    folder: str = Query(...),
    filename: str = Query(...),
) -> FileResponse:
    carpeta = OUTPUT_CONSULTAS / folder
    archivo = carpeta / filename
    try:
        archivo.resolve().relative_to(OUTPUT_CONSULTAS.resolve())
    except ValueError:
        raise HTTPException(status_code=400, detail="Ruta no permitida.")
    if not archivo.exists() or not archivo.is_file():
        raise HTTPException(status_code=404, detail="Archivo no encontrado.")
    return FileResponse(
        path=str(archivo), filename=filename, media_type="application/pdf"
    )

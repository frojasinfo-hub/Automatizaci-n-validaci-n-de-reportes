"""Genera y actualiza el reporte consolidado Excel."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import List

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.infrastructure.logger_manager import LoggerManager
from src.models.tercero import EstadoFuente, ResultadoTercero

logger = LoggerManager()

_COLORES_ESTADO: dict[str, str] = {
    "EXITOSO": "C6EFCE",
    "AUTOMATICO": "C6EFCE",
    "PARCIAL": "FFEB9C",
    "EN_ESPERA": "FFEB9C",
    "NO_ENCONTRADO": "FCE4D6",   # Naranja claro — consultado, sin registro
    "FALLIDO": "FFC7CE",
    "FALLIDO_TOTAL": "FFC7CE",
    "ERROR_PDF": "FFC7CE",
    "PENDIENTE": "D9D9D9",
}

_ENCABEZADOS = [
    "Cédula", "Tipo Doc", "Nombre", "Tipo Persona",
    "Procuraduría", "Hallazgo Procuraduría",
    "Policía",      "Hallazgo Policía",
    "Fiscal",       "Hallazgo Fiscal",
    "Estado",
    "Fecha", "Número de intentos", "Fecha segunda validación", "Observacion",
    "NIT Inspektor", "Dígito verificación", "NIT completo",
]

# Índices de columna (1-based)
_C_CEDULA = 1
_C_TIPO_DOC = 2
_C_NOMBRE = 3
_C_TIPO_PERSONA = 4
_C_PRO = 5
_C_HALL_PRO = 6
_C_POL = 7
_C_HALL_POL = 8
_C_FIS = 9
_C_HALL_FIS = 10
_C_ESTADO = 11
_C_FECHA_PRIMERA = 12
_C_CONSULTAS = 13
_C_ULTIMA_CONSULTA = 14
_C_ERROR = 15
_C_NIT_BASE = 16
_C_NIT_DV = 17
_C_NIT_COMPLETO = 18

_ANCHOS_COL = [15, 10, 40, 18, 15, 55, 12, 55, 12, 55, 15, 12, 10, 18, 50, 12, 20, 15]


class ReporteExcel:
    """Mantiene resultado_consultas.xlsx con historial acumulado por ejecución."""

    def cargar_estados(self, path: Path) -> dict[str, dict]:
        """Lee el Excel existente y retorna {cedula: {estado, pro, pol, fis, consultas, fecha_primera}}."""
        if not path.exists():
            return {}
        try:
            wb = openpyxl.load_workbook(path)
            ws = wb.active
            result: dict[str, dict] = {}
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[_C_CEDULA - 1]:
                    continue
                cedula = str(row[_C_CEDULA - 1]).strip()
                if cedula in ("None", ""):
                    continue

                def _g(idx: int, default=None, r=row):
                    return r[idx] if len(r) > idx else default

                result[cedula] = {
                    "estado": str(_g(_C_ESTADO - 1) or "DESCONOCIDO"),
                    "procuraduria": str(_g(_C_PRO - 1) or "PENDIENTE"),
                    "policia": str(_g(_C_POL - 1) or "PENDIENTE"),
                    "fiscal": str(_g(_C_FIS - 1) or "PENDIENTE"),
                    "consultas": int(_g(_C_CONSULTAS - 1) or 0),
                    "fecha_primera": _g(_C_FECHA_PRIMERA - 1),
                }
            return result
        except Exception as exc:
            logger.warning(f"[Excel] No se pudo leer el archivo existente: {exc}")
            return {}

    def actualizar_o_crear(
        self,
        resultados: List[ResultadoTercero],
        path: Path,
        incrementar_consulta: bool = True,
    ) -> Path:
        """Actualiza filas existentes o agrega nuevas preservando el historial.

        incrementar_consulta=False se usa para reintentos dentro de la misma ejecución
        para no contar el mismo run más de una vez.
        """
        path.parent.mkdir(parents=True, exist_ok=True)
        estados_previos = self.cargar_estados(path)

        if path.exists():
            try:
                wb = openpyxl.load_workbook(path)
                ws = wb.active
            except Exception:
                wb, ws = self._nuevo_libro()
        else:
            wb, ws = self._nuevo_libro()

        # Mapa {cedula: numero_fila} del archivo actual
        cedula_a_fila: dict[str, int] = {}
        for row_idx in range(2, ws.max_row + 1):
            val = ws.cell(row_idx, _C_CEDULA).value
            if val:
                cedula_a_fila[str(val).strip()] = row_idx

        ahora = datetime.now()
        for resultado in resultados:
            cedula = resultado.tercero.numero_documento
            previo = estados_previos.get(cedula, {})
            consultas_previas = previo.get("consultas", 0)
            fecha_primera = previo.get("fecha_primera") or ahora.strftime("%d/%m/%y")

            consultas = (consultas_previas + 1) if incrementar_consulta else consultas_previas

            if cedula in cedula_a_fila:
                fila = cedula_a_fila[cedula]
            else:
                fila = ws.max_row + 1
                cedula_a_fila[cedula] = fila

            self._escribir_fila(
                ws, fila, resultado,
                consultas=consultas,
                fecha_primera=str(fecha_primera),
                ultima_consulta=ahora.strftime("%d/%m/%y %H:%M"),
            )

        self._ajustar_columnas(ws)
        self._guardar_con_reintento(wb, path)
        logger.info(f"[Excel] Reporte actualizado: {path}")
        return path

    def generar(self, resultados: List[ResultadoTercero], output_path: Path) -> Path:
        """Backward compatibility — llama a actualizar_o_crear."""
        return self.actualizar_o_crear(resultados, output_path, incrementar_consulta=True)

    # --- Helpers internos ---

    def _nuevo_libro(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Consultas"
        self._escribir_encabezados(ws)
        return wb, ws

    def _escribir_encabezados(self, ws) -> None:
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(fill_type="solid", fgColor="1F3864")
        for col, texto in enumerate(_ENCABEZADOS, start=1):
            celda = ws.cell(row=1, column=col, value=texto)
            celda.font = header_font
            celda.fill = header_fill
            celda.alignment = Alignment(horizontal="center")

    _TERMINADOS_FUENTE = frozenset({
        EstadoFuente.EXITOSO, EstadoFuente.AUTOMATICO, EstadoFuente.NO_ENCONTRADO
    })
    _EXITOSOS_FUENTE = frozenset({EstadoFuente.EXITOSO, EstadoFuente.AUTOMATICO})

    def _calcular_estado(self, resultado: ResultadoTercero) -> str:
        """Estado general del tercero.

        Para CC: NO_ENCONTRADO cuenta como terminal (persona sin antecedentes = OK).
        Para NIT: solo cuenta EXITOSO/AUTOMATICO (se requiere descarga real).
                  NO_ENCONTRADO en NIT = PARCIAL (empresa sin certificado descargado).
        """
        es_nit = resultado.tercero.tipo_documento == "NIT"
        fuentes = (
            [resultado.procuraduria, resultado.fiscal]
            if es_nit
            else [resultado.procuraduria, resultado.policia, resultado.fiscal]
        )
        total = len(fuentes)

        if es_nit:
            # Procuraduria: EXITOSO, AUTOMATICO o NO_ENCONTRADO = terminal válido
            pro_ok = resultado.procuraduria.estado in self._EXITOSOS_FUENTE or \
                     resultado.procuraduria.estado == EstadoFuente.NO_ENCONTRADO
            # Fiscal: solo EXITOSO o AUTOMATICO
            fis_ok = resultado.fiscal.estado in self._EXITOSOS_FUENTE
            return "EXITOSO" if (pro_ok and fis_ok) else "PARCIAL"

        # Para CC
        terminadas = sum(1 for f in fuentes if f.estado in self._TERMINADOS_FUENTE)
        if terminadas == total:
            return "EXITOSO"
        if terminadas == 0:
            return "FALLIDO_TOTAL"
        return "PARCIAL"

    def _errores_detalle(self, resultado: ResultadoTercero) -> str:
        partes = []
        for prefijo, fuente in [
            ("PRO", resultado.procuraduria),
            ("POL", resultado.policia),
            ("FIS", resultado.fiscal),
        ]:
            if fuente.error_mensaje:
                partes.append(f"{prefijo}: {fuente.error_mensaje[:80]}")
        return " | ".join(partes)

    def _escribir_fila(
        self,
        ws,
        fila: int,
        resultado: ResultadoTercero,
        consultas: int,
        fecha_primera: str,
        ultima_consulta: str,
    ) -> None:
        es_nit = resultado.tercero.tipo_documento == "NIT"
        tipo_persona = "Persona Jurídica" if es_nit else "Persona Natural"
        estado_general = self._calcular_estado(resultado)

        # Desglose NIT: solo para Persona Jurídica con NIT de 10 dígitos
        nd = resultado.tercero.numero_documento
        nit_base: object = None
        nit_dv: object = None
        nit_completo: object = None
        if es_nit and len(nd) == 10 and nd.isdigit():
            nit_base = nd[:9]
            nit_dv = nd[9]
            nit_completo = nd

        datos_estado: dict[int, object] = {
            _C_CEDULA: nd,
            _C_TIPO_DOC: resultado.tercero.tipo_documento,
            _C_NOMBRE: resultado.tercero.nombre_completo,
            _C_TIPO_PERSONA: tipo_persona,
            _C_PRO: resultado.procuraduria.estado.value,
            _C_POL: "NO APLICA" if es_nit else resultado.policia.estado.value,
            _C_FIS: resultado.fiscal.estado.value,
            _C_ESTADO: estado_general,
            _C_FECHA_PRIMERA: fecha_primera,
            _C_CONSULTAS: consultas,
            _C_ULTIMA_CONSULTA: ultima_consulta,
            _C_ERROR: self._errores_detalle(resultado),
            _C_NIT_BASE: nit_base,
            _C_NIT_DV: nit_dv,
            _C_NIT_COMPLETO: nit_completo,
        }

        datos_hallazgo: dict[int, str] = {
            _C_HALL_PRO: resultado.procuraduria.hallazgo,
            _C_HALL_POL: "NO APLICA" if es_nit else resultado.policia.hallazgo,
            _C_HALL_FIS: resultado.fiscal.hallazgo,
        }

        color = _COLORES_ESTADO.get(estado_general, "FFFFFF")
        fill_estado = PatternFill(fill_type="solid", fgColor=color)
        fill_hallazgo = PatternFill(fill_type="solid", fgColor="FFFFFF")

        for col, valor in datos_estado.items():
            celda = ws.cell(row=fila, column=col, value=valor)
            celda.alignment = Alignment(horizontal="center")
            celda.fill = fill_estado

        for col, valor in datos_hallazgo.items():
            celda = ws.cell(row=fila, column=col, value=valor)
            celda.alignment = Alignment(horizontal="left", wrap_text=True)
            celda.fill = fill_hallazgo

    def _ajustar_columnas(self, ws) -> None:
        for idx, ancho in enumerate(_ANCHOS_COL, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = ancho

    def _guardar_con_reintento(self, wb, path: Path) -> None:
        try:
            wb.save(path)
        except PermissionError:
            print(
                f"\n{'='*60}\n"
                f"  ERROR: No se pudo guardar '{path.name}'.\n"
                f"  Cierre el archivo en Excel y presione ENTER...\n"
                f"{'='*60}"
            )
            input()
            wb.save(path)
